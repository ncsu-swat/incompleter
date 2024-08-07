#Source: https://stackoverflow.com/questions/61278770/attributeerror-str-object-has-no-attribute-id-with-openpyxl
import openpyxl
from openpyxl.chart import (PieChart, Reference)

#Load an existing workbook to manipulate
workbook_object = openpyxl.load_workbook('Mitchell Bridge Rd.xlsx')

#Rename the first sheet as 'Originale data'
worksheet_object0 = workbook_object.active
worksheet_object0.title = 'Original Data'

#Create a new worksheet object and call in 'Filtered data
worksheet_object1 = workbook_object.create_sheet('Filtered Data')

#Copy 'Original Data' to 'Filtered Data' worksheet
for i in range(1, worksheet_object0.max_row + 1):
    for j in range(1, worksheet_object0.max_column + 1):
        worksheet_object1.cell(row=i, column=j).value = worksheet_object0.cell(row=i, column=j).value

#Remove all the unwanted columns from 'Filtered Data' worksheet
unwanted_columns = ['AccidentNumber', 'County', 'RouteType', 'Route', 'Milelog', 'IntersectingRoute',
                  'RampSection', 'DistanceFrom', 'DirectionFrom', 'LocationOfImpact', 'FirstHarmfulEvent',
                  'DirVeh1', 'DirVeh2', 'MnvrVeh1', 'MnvrVeh2', 'MicrofilmNo', 'U1Factors', 'U2Factors',
                  'Vendor', 'IntersectRouteType', 'U1FirstHarmfulEvent', 'U2FirstHarmfulEvent',
                  ]

#delete unwated columns
for i in range(1, worksheet_object1.max_column + 1):
    for column in unwanted_columns:
        if worksheet_object1.cell(row=1, column=i).value == column:
            worksheet_object1.delete_cols(i)

#Store crash data into dictionary
crash_data = {'Angle': 0,
              'Side-Swipe': 0,
              'Rear End': 0,
              'Head On': 0,
              }

##Populate dictionary with data from worksheet
for i in range(2, worksheet_object1.max_row + 1):
    manner_of_collision = worksheet_object1.cell(row=i, column=6).value
    if manner_of_collision in crash_data:
        crash_data[manner_of_collision] += 1

#Convert data into a list
data = [['Manner of Collision', 'Number of Crash']]
for key, value in crash_data.items():
    data.append([key, value])

#New worksheet for Pie chart
worksheet_object2 = workbook_object.create_sheet('Pie Chart')
for row in data:
    worksheet_object2.append(row)

#Pie Chart plot
pie = PieChart()
labels = Reference(worksheet_object2, min_col=1, min_row=2, max_row=5)
data = Reference(worksheet_object2, min_col=2, min_row=1, max_row=5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.title = 'Manner of Collision'

worksheet_object2.add_chart('D1')

#Save the manipulated workbook
workbook_object.save('pie.xlsx')