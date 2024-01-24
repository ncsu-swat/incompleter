# LExecutor: DO NOT INSTRUMENT

#Source: https://stackoverflow.com/questions/48516372/getting-typeerror-during-creation-of-bar-chart-from-excel-using-openpyxl-referen
#!/usr/bin/python

from lexecutor.Runtime import _n_
from lexecutor.Runtime import _a_
from lexecutor.Runtime import _c_
from lexecutor.Runtime import _l_
try:
    from openpyxl import load_workbook
    _l_(532782)

except ImportError:
    pass
try:
    from openpyxl import Workbook
    _l_(532784)

except ImportError:
    pass
try:
    from openpyxl.chart import (
     Reference,
     Series,
     BarChart
    )
    _l_(532786)

except ImportError:
    pass

wb = _c_(532788, _n_(532787, "load_workbook", lambda: load_workbook), 'revenue.xlsx')
_l_(532789)
wsheet = _c_(532792, _a_(532791, _n_(532790, "wb", lambda: wb), "get_sheet_by_name"), 'sales')
_l_(532793)
_c_(532796, _n_(532794, "print", lambda: print), _n_(532795, "wsheet", lambda: wsheet))
_l_(532797)


data =   _c_(532800, _n_(532798, "Reference", lambda: Reference), _n_(532799, "wsheet", lambda: wsheet), (5, 2), (5, 10))
_l_(532801)
categs = _c_(532804, _n_(532802, "Reference", lambda: Reference), _n_(532803, "wsheet", lambda: wsheet), (3, 2), (3, 10))
_l_(532805)

chart = _c_(532807, _n_(532806, "BarChart", lambda: BarChart))
_l_(532808)
_c_(532812, _a_(532810, _n_(532809, "chart", lambda: chart), "add_data"), data=_n_(532811, "data", lambda: data))
_l_(532813)
_c_(532817, _a_(532815, _n_(532814, "chart", lambda: chart), "set_categories"), _n_(532816, "categs", lambda: categs))
_l_(532818)

_n_(532819, "chart", lambda: chart).legend = None
_l_(532820)
_a_(532822, _n_(532821, "chart", lambda: chart), "y_axis").majorGridlines = None
_l_(532823)
_n_(532824, "chart", lambda: chart).varyColors = True
_l_(532825)
_n_(532826, "chart", lambda: chart).title = "Sales By name"
_l_(532827)

_c_(532831, _a_(532829, _n_(532828, "wsheet", lambda: wsheet), "add_chart"), _n_(532830, "chart", lambda: chart), "H15")    
_l_(532832)    

_c_(532835, _a_(532834, _n_(532833, "wb", lambda: wb), "save"), "revenue.xlsx")
_l_(532836)

_c_(532838, _n_(532837, "print", lambda: print), "end")
_l_(532839)