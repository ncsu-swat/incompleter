# LExecutor: DO NOT INSTRUMENT

#Source: https://stackoverflow.com/questions/39077740/copying-a-range-of-cells-from-one-workbook-to-another-with-openpyxl-attributee
from lexecutor.Runtime import _n_
from lexecutor.Runtime import _a_
from lexecutor.Runtime import _c_
from lexecutor.Runtime import _l_
try:
    from openpyxl import Workbook, load_workbook
    _l_(542875)

except ImportError:
    pass


wb_raw_data = _c_(542877, _n_(542876, "load_workbook", lambda: load_workbook), filename = 'D:\Folder\Test.xlsx')
_l_(542878)

sheet1 = _c_(542881, _a_(542880, _n_(542879, "wb_raw_data", lambda: wb_raw_data), "get_sheet_by_name"), 'Sheet1')
_l_(542882)

wb_to_change_data = _c_(542884, _n_(542883, "load_workbook", lambda: load_workbook), filename = 'D:\Folder\Existing.xlsx')
_l_(542885)

sheet2 = _c_(542888, _a_(542887, _n_(542886, "wb_to_change_data", lambda: wb_to_change_data), "get_sheet_by_name"), 'Sheet4')
_l_(542889)

cellrange = _a_(542891, _n_(542890, "sheet1", lambda: sheet1)['A1':'B20'], "value")
_l_(542892)

_n_(542893, "sheet2", lambda: sheet2)['A1':'B20'] = _n_(542894, "cellrange", lambda: cellrange)
_l_(542895)

_c_(542898, _a_(542897, _n_(542896, "wb_to_change_data", lambda: wb_to_change_data), "save"), 'D:\Folder\Existing.xlsx')
_l_(542899)