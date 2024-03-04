#Source: https://stackoverflow.com/questions/59241010/how-to-fix-typeerror-cannot-serialize-io-bufferedreader-object-when-writin
# Path to .xlsx
MasterItem = MonthlyFolder + "MasterItem__Nov2019.xlsx"

# Function to read the excel file
def ReadExcel(filename, sheetname=None, header=0):
    from openpyxl import load_workbook

    wb = load_workbook(filename, read_only=True)

    if sheetname is None:  # If sheetname is not provided then grab the first sheet
        print("\t Reading " + wb.sheetnames[0])
        ws = wb[wb.sheetnames[0]]
    else:
        print("\t Reading " + sheetname)
        ws = wb[sheetname]

    data = ws.values

    if header is None:
        columns = None
    elif header > 0:
        # Skip non header rows
        for i in range(0, header):
            next(data)
        # Save header row
        columns = next(data)[0:]
    else:
        columns = next(data)[0:]

    # Create a DataFrame based on the subsequent lines of data
    df_Out = pd.DataFrame(data, columns=columns)

    return df_Out

# Reading .xlsx and writing as pickle
RawMasterItem = ReadExcel(MasterItem)
pd.to_pickle(RawMasterItem, MonthlyFolder+"RawMasterItem.pkl") # This fails to run