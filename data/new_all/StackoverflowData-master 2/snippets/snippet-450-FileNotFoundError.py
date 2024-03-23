#Source: https://stackoverflow.com/questions/55918587/openpyxl-filenotfounderror-errno-2-no-such-file-or-directory
from os import listdir
from openpyxl import load_workbook, Workbook

files = listdir(r'C:\Users\MyID\Desktop\auto_task\examples\06')
result_xlsx = Workbook()
result_sheet =result_xlsx.active

for file in files:
    if file[-4:] != 'xlsx':
        continue

    tg_xlsx = load_workbook(file, read_only=True)
    tg_sheet = tg_xlsx.active

    for row in tg_sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        result_sheet.append(row_data)

result_xlsx.save('result.xlsx')