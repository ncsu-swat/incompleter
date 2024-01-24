#Source: https://stackoverflow.com/questions/39077740/copying-a-range-of-cells-from-one-workbook-to-another-with-openpyxl-attributee
from openpyxl import Workbook, load_workbook


wb_raw_data = load_workbook(filename = 'D:\Folder\Test.xlsx')

sheet1 = wb_raw_data.get_sheet_by_name('Sheet1')

wb_to_change_data = load_workbook(filename = 'D:\Folder\Existing.xlsx')

sheet2 = wb_to_change_data.get_sheet_by_name('Sheet4')

cellrange = sheet1['A1':'B20'].value

sheet2['A1':'B20'] = cellrange

wb_to_change_data.save('D:\Folder\Existing.xlsx')