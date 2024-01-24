#Source: https://stackoverflow.com/questions/75709003/attributeerror-module-emotions-has-no-attribute-emotion
from openpyxl import workbook, load_workbook
import random
import main


text = main.HomeAI().text_given

def not_found():
    notfound = "Well I have not seen this emotion before so I will try my hardest to understand"
    return notfound

def check():
    return msg

def goodBad():
    emotion = gb
    return emotion

def look():
    global msg
    count = 0
    global gb
    file = "emotions.xlsx"
    wb = load_workbook(file)
    ws = wb.active
    while True:
        count = count + 1
        if count == ws.max_row:
            msg = '0004302X'
            check()

        A = (ws[f'A{count}'].value)
        if A == text:
            msg = '00492X'
            gb = (ws[f'B{count}'].value)

        else:
            continue