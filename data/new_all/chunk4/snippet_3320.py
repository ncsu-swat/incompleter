# LExecutor: DO NOT INSTRUMENT

#Source: https://stackoverflow.com/questions/75709003/attributeerror-module-emotions-has-no-attribute-emotion
from lexecutor.Runtime import _n_
from lexecutor.Runtime import _a_
from lexecutor.Runtime import _c_
from lexecutor.Runtime import _l_
try:
    from openpyxl import workbook, load_workbook
    _l_(619625)

except ImportError:
    pass
try:
    import random
    _l_(619627)

except ImportError:
    pass
try:
    import main
    _l_(619629)

except ImportError:
    pass


text = _a_(619633, _c_(619632, _a_(619631, _n_(619630, "main", lambda: main), "HomeAI")), "text_given")
_l_(619634)

def not_found():
    _l_(619638)

    notfound = "Well I have not seen this emotion before so I will try my hardest to understand"
    _l_(619635)
    aux = _n_(619636, "notfound", lambda: notfound)
    _l_(619637)
    return aux

def check():
    _l_(619641)

    aux = _n_(619639, "msg", lambda: msg)
    _l_(619640)
    return aux

def goodBad():
    _l_(619646)

    emotion = _n_(619642, "gb", lambda: gb)
    _l_(619643)
    aux = _n_(619644, "emotion", lambda: emotion)
    _l_(619645)
    return aux

def look():
    _l_(619682)

    global msg
    _l_(619647)
    count = 0
    _l_(619648)
    global gb
    _l_(619649)
    file = "emotions.xlsx"
    _l_(619650)
    wb = _c_(619653, _n_(619651, "load_workbook", lambda: load_workbook), _n_(619652, "file", lambda: file))
    _l_(619654)
    ws = _a_(619656, _n_(619655, "wb", lambda: wb), "active")
    _l_(619657)
    while True:
        _l_(619681)

        count = _n_(619658, "count", lambda: count) + 1
        _l_(619659)
        if _n_(619660, "count", lambda: count) == _a_(619662, _n_(619661, "ws", lambda: ws), "max_row"):
            _l_(619667)

            msg = '0004302X'
            _l_(619663)
            _c_(619665, _n_(619664, "check", lambda: check))
            _l_(619666)

        A = _a_(619670, _n_(619668, "ws", lambda: ws)[f'A{_n_(619669, "count", lambda: count)}'], 'value')
        _l_(619671)
        if _n_(619672, 'A', lambda: A) == _n_(619673, 'text', lambda: text):
            _l_(619680)

            msg = '00492X'
            _l_(619674)
            gb = _a_(619677, _n_(619675, 'ws', lambda: ws)[f'B{_n_(619676, "count", lambda: count)}'], 'value')
            _l_(619678)

        else:
            continue
            _l_(619679)