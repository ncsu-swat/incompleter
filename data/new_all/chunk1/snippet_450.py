# LExecutor: DO NOT INSTRUMENT

#Source: https://stackoverflow.com/questions/55918587/openpyxl-filenotfounderror-errno-2-no-such-file-or-directory
from lexecutor.Runtime import _n_
from lexecutor.Runtime import _a_
from lexecutor.Runtime import _c_
from lexecutor.Runtime import _l_
try:
    from os import listdir
    _l_(392387)

except ImportError:
    pass
try:
    from openpyxl import load_workbook, Workbook
    _l_(392389)

except ImportError:
    pass

files = _c_(392391, _n_(392390, "listdir", lambda: listdir), r'C:\Users\MyID\Desktop\auto_task\examples\06')
_l_(392392)
result_xlsx = _c_(392394, _n_(392393, "Workbook", lambda: Workbook))
_l_(392395)
result_sheet =_a_(392397, _n_(392396, "result_xlsx", lambda: result_xlsx), "active")
_l_(392398)

for file in _n_(392399, "files", lambda: files):
    _l_(392428)

    if _n_(392400, "file", lambda: file)[-4:] != 'xlsx':
        _l_(392402)

        continue
        _l_(392401)

    tg_xlsx = _c_(392405, _n_(392403, "load_workbook", lambda: load_workbook), _n_(392404, "file", lambda: file), read_only=True)
    _l_(392406)
    tg_sheet = _a_(392408, _n_(392407, "tg_xlsx", lambda: tg_xlsx), "active")
    _l_(392409)

    for row in _c_(392412, _a_(392411, _n_(392410, "tg_sheet", lambda: tg_sheet), "iter_rows")):
        _l_(392427)

        row_data = []
        _l_(392413)
        for cell in _n_(392414, "row", lambda: row):
            _l_(392421)

            _c_(392419, _a_(392416, _n_(392415, "row_data", lambda: row_data), "append"), _a_(392418, _n_(392417, "cell", lambda: cell), "value"))
            _l_(392420)
        _c_(392425, _a_(392423, _n_(392422, "result_sheet", lambda: result_sheet), "append"), _n_(392424, "row_data", lambda: row_data))
        _l_(392426)

_c_(392431, _a_(392430, _n_(392429, "result_xlsx", lambda: result_xlsx), "save"), 'result.xlsx')
_l_(392432)