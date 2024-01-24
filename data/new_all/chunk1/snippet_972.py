# LExecutor: DO NOT INSTRUMENT

#Source: https://stackoverflow.com/questions/61278770/attributeerror-str-object-has-no-attribute-id-with-openpyxl
from lexecutor.Runtime import _n_
from lexecutor.Runtime import _a_
from lexecutor.Runtime import _c_
from lexecutor.Runtime import _l_
try:
    import openpyxl
    _l_(397758)

except ImportError:
    pass
try:
    from openpyxl.chart import (PieChart, Reference)
    _l_(397760)

except ImportError:
    pass

#Load an existing workbook to manipulate
workbook_object = _c_(397763, _a_(397762, _n_(397761, "openpyxl", lambda: openpyxl), "load_workbook"), 'Mitchell Bridge Rd.xlsx')
_l_(397764)

#Rename the first sheet as 'Originale data'
worksheet_object0 = _a_(397766, _n_(397765, "workbook_object", lambda: workbook_object), "active")
_l_(397767)
_n_(397768, "worksheet_object0", lambda: worksheet_object0).title = 'Original Data'
_l_(397769)

#Create a new worksheet object and call in 'Filtered data
worksheet_object1 = _c_(397772, _a_(397771, _n_(397770, "workbook_object", lambda: workbook_object), "create_sheet"), 'Filtered Data')
_l_(397773)

#Copy 'Original Data' to 'Filtered Data' worksheet
for i in _c_(397777, _n_(397774, "range", lambda: range), 1, _a_(397776, _n_(397775, "worksheet_object0", lambda: worksheet_object0), "max_row") + 1):
    _l_(397795)

    for j in _c_(397781, _n_(397778, "range", lambda: range), 1, _a_(397780, _n_(397779, "worksheet_object0", lambda: worksheet_object0), "max_column") + 1):
        _l_(397794)

        _c_(397786, _a_(397783, _n_(397782, "worksheet_object1", lambda: worksheet_object1), "cell"), row=_n_(397784, "i", lambda: i), column=_n_(397785, "j", lambda: j)).value = _a_(397792, _c_(397791, _a_(397788, _n_(397787, "worksheet_object0", lambda: worksheet_object0), "cell"), row=_n_(397789, "i", lambda: i), column=_n_(397790, "j", lambda: j)), "value")
        _l_(397793)

#Remove all the unwanted columns from 'Filtered Data' worksheet
unwanted_columns = ['AccidentNumber', 'County', 'RouteType', 'Route', 'Milelog', 'IntersectingRoute',
                  'RampSection', 'DistanceFrom', 'DirectionFrom', 'LocationOfImpact', 'FirstHarmfulEvent',
                  'DirVeh1', 'DirVeh2', 'MnvrVeh1', 'MnvrVeh2', 'MicrofilmNo', 'U1Factors', 'U2Factors',
                  'Vendor', 'IntersectRouteType', 'U1FirstHarmfulEvent', 'U2FirstHarmfulEvent',
                  ]
_l_(397796)

#delete unwated columns
for i in _c_(397800, _n_(397797, "range", lambda: range), 1, _a_(397799, _n_(397798, "worksheet_object1", lambda: worksheet_object1), "max_column") + 1):
    _l_(397815)

    for column in _n_(397801, "unwanted_columns", lambda: unwanted_columns):
        _l_(397814)

        if _a_(397806, _c_(397805, _a_(397803, _n_(397802, "worksheet_object1", lambda: worksheet_object1), "cell"), row=1, column=_n_(397804, "i", lambda: i)), "value") == _n_(397807, "column", lambda: column):
            _l_(397813)

            _c_(397811, _a_(397809, _n_(397808, "worksheet_object1", lambda: worksheet_object1), "delete_cols"), _n_(397810, "i", lambda: i))
            _l_(397812)

#Store crash data into dictionary
crash_data = {'Angle': 0,
              'Side-Swipe': 0,
              'Rear End': 0,
              'Head On': 0,
              }
_l_(397816)

##Populate dictionary with data from worksheet
for i in _c_(397820, _n_(397817, "range", lambda: range), 2, _a_(397819, _n_(397818, "worksheet_object1", lambda: worksheet_object1), "max_row") + 1):
    _l_(397833)

    manner_of_collision = _a_(397825, _c_(397824, _a_(397822, _n_(397821, "worksheet_object1", lambda: worksheet_object1), "cell"), row=_n_(397823, "i", lambda: i), column=6), "value")
    _l_(397826)
    if _n_(397827, "manner_of_collision", lambda: manner_of_collision) in _n_(397828, "crash_data", lambda: crash_data):
        _l_(397832)

        _n_(397829, "crash_data", lambda: crash_data)[_n_(397830, "manner_of_collision", lambda: manner_of_collision)] += 1
        _l_(397831)

#Convert data into a list
data = [['Manner of Collision', 'Number of Crash']]
_l_(397834)
for key, value in _c_(397837, _a_(397836, _n_(397835, "crash_data", lambda: crash_data), "items")):
    _l_(397844)

    _c_(397842, _a_(397839, _n_(397838, "data", lambda: data), "append"), [_n_(397840, "key", lambda: key), _n_(397841, "value", lambda: value)])
    _l_(397843)

#New worksheet for Pie chart
worksheet_object2 = _c_(397847, _a_(397846, _n_(397845, "workbook_object", lambda: workbook_object), "create_sheet"), 'Pie Chart')
_l_(397848)
for row in _n_(397849, "data", lambda: data):
    _l_(397855)

    _c_(397853, _a_(397851, _n_(397850, "worksheet_object2", lambda: worksheet_object2), "append"), _n_(397852, "row", lambda: row))
    _l_(397854)

#Pie Chart plot
pie = _c_(397857, _n_(397856, "PieChart", lambda: PieChart))
_l_(397858)
labels = _c_(397861, _n_(397859, "Reference", lambda: Reference), _n_(397860, "worksheet_object2", lambda: worksheet_object2), min_col=1, min_row=2, max_row=5)
_l_(397862)
data = _c_(397865, _n_(397863, "Reference", lambda: Reference), _n_(397864, "worksheet_object2", lambda: worksheet_object2), min_col=2, min_row=1, max_row=5)
_l_(397866)
_c_(397870, _a_(397868, _n_(397867, "pie", lambda: pie), "add_data"), _n_(397869, "data", lambda: data), titles_from_data=True)
_l_(397871)
_c_(397875, _a_(397873, _n_(397872, "pie", lambda: pie), "set_categories"), _n_(397874, "labels", lambda: labels))
_l_(397876)
_n_(397877, "pie", lambda: pie).title = 'Manner of Collision'
_l_(397878)

_c_(397881, _a_(397880, _n_(397879, "worksheet_object2", lambda: worksheet_object2), "add_chart"), 'D1')
_l_(397882)

#Save the manipulated workbook
_c_(397885, _a_(397884, _n_(397883, "workbook_object", lambda: workbook_object), "save"), 'pie.xlsx')
_l_(397886)