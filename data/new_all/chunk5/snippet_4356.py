# LExecutor: DO NOT INSTRUMENT

#Source: https://stackoverflow.com/questions/59241010/how-to-fix-typeerror-cannot-serialize-io-bufferedreader-object-when-writin
# Path to .xlsx
from lexecutor.Runtime import _n_
from lexecutor.Runtime import _a_
from lexecutor.Runtime import _c_
from lexecutor.Runtime import _l_
MasterItem = _n_(647559, "MonthlyFolder", lambda: MonthlyFolder) + "MasterItem__Nov2019.xlsx"
_l_(647560)

# Function to read the excel file
def ReadExcel(filename, sheetname=None, header=0):
    _l_(647617)

    try:
        from openpyxl import load_workbook
        _l_(647562)

    except ImportError:
        pass

    wb = _c_(647565, _n_(647563, "load_workbook", lambda: load_workbook), _n_(647564, "filename", lambda: filename), read_only=True)
    _l_(647566)

    if _n_(647567, "sheetname", lambda: sheetname) is None:
        _l_(647584)

        _c_(647571, _n_(647568, "print", lambda: print), "\t Reading " + _a_(647570, _n_(647569, "wb", lambda: wb), "sheetnames")[0])
        _l_(647572)
        ws = _n_(647573, "wb", lambda: wb)[_a_(647575, _n_(647574, "wb", lambda: wb), "sheetnames")[0]]
        _l_(647576)
    else:
        _c_(647579, _n_(647577, "print", lambda: print), "\t Reading " + _n_(647578, "sheetname", lambda: sheetname))
        _l_(647580)
        ws = _n_(647581, "wb", lambda: wb)[_n_(647582, "sheetname", lambda: sheetname)]
        _l_(647583)

    data = _a_(647586, _n_(647585, "ws", lambda: ws), "values")
    _l_(647587)

    if _n_(647588, "header", lambda: header) is None:
        _l_(647608)

        columns = None
        _l_(647589)
    elif _n_(647590, "header", lambda: header) > 0:
        _l_(647607)

        # Skip non header rows
        for i in _c_(647593, _n_(647591, "range", lambda: range), 0, _n_(647592, "header", lambda: header)):
            _l_(647598)

            _c_(647596, _n_(647594, "next", lambda: next), _n_(647595, "data", lambda: data))
            _l_(647597)
        # Save header row
        columns = _c_(647601, _n_(647599, "next", lambda: next), _n_(647600, "data", lambda: data))[0:]
        _l_(647602)
    else:
        columns = _c_(647605, _n_(647603, "next", lambda: next), _n_(647604, "data", lambda: data))[0:]
        _l_(647606)

    # Create a DataFrame based on the subsequent lines of data
    df_Out = _c_(647613, _a_(647610, _n_(647609, "pd", lambda: pd), "DataFrame"), _n_(647611, "data", lambda: data), columns=_n_(647612, "columns", lambda: columns))
    _l_(647614)
    aux = _n_(647615, "df_Out", lambda: df_Out)
    _l_(647616)

    return aux

# Reading .xlsx and writing as pickle
RawMasterItem = _c_(647620, _n_(647618, "ReadExcel", lambda: ReadExcel), _n_(647619, "MasterItem", lambda: MasterItem))
_l_(647621)
_c_(647626, _a_(647623, _n_(647622, "pd", lambda: pd), "to_pickle"), _n_(647624, "RawMasterItem", lambda: RawMasterItem), _n_(647625, "MonthlyFolder", lambda: MonthlyFolder)+"RawMasterItem.pkl") # This fails to run
_l_(647627) # This fails to run