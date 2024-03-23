# Import the required modules
import datetime
from openpyxl import Workbook
import time
  
  
# Main Function
if __name__ == '__main__':
    
    # Create a worbook object
    wb = Workbook()
  
    # Select the active sheet
    ws = wb.active
  
    # Heading of Cell A1
    ws.cell(row=1, column=1).value = "Current Date and Time"
  
    # Cell A2 containing the Current Date and Time
    ws.cell(row=2, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
  
    # Sleep of 2 seconds
    time.sleep(2)
  
    # Cell A3 containing the Current Date and Time
    ws.cell(row=3, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    time.sleep(2)
  
    # Cell A4 containing the Current Date and Time
    ws.cell(row=4, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
  
    # Save the workbook with a
    # filename and close the object
    wb.save('gfg.xlsx')
    wb.close()