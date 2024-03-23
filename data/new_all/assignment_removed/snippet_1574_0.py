import datetime
from openpyxl import Workbook
import time
if __name__ == '__main__':
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = 'Current Date and Time'
    ws.cell(row=2, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    time.sleep(2)
    ws.cell(row=3, column=1).value = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    time.sleep(2)
    wb.save('gfg.xlsx')
    wb.close()